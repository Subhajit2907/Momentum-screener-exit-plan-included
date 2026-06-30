"""
Core screener logic — callable as a library by app.py.
No sys.exit(), no print() — returns data or raises exceptions.
"""

import io
import os
from datetime import datetime

import pandas as pd
import yfinance as yf
from openpyxl.styles import Alignment, Font, PatternFill

# ── Filter thresholds ──────────────────────────────────────
RSI_MIN        = 50
MIN_AVG_VOLUME = 500_000
NEAR_HIGH_PCT  = 20

# ── Composite score weights (must sum to 1.0) ──────────────
W_EMA200 = 0.30
W_RSI    = 0.25
W_52W    = 0.25
W_VOLUME = 0.20


# ── Indicators ────────────────────────────────────────────

def calc_ema(prices, period):
    if len(prices) < period:
        return None
    k = 2 / (period + 1)
    ema = float(sum(prices[:period]) / period)
    for price in prices[period:]:
        ema = float(price) * k + ema * (1 - k)
    return ema


def calc_rsi(prices, period=14):
    if len(prices) < period + 1:
        return None
    deltas = [prices[i] - prices[i - 1] for i in range(1, len(prices))]
    gains  = [d if d > 0 else 0.0 for d in deltas]
    losses = [-d if d < 0 else 0.0 for d in deltas]
    avg_gain = sum(gains[:period]) / period
    avg_loss = sum(losses[:period]) / period
    for i in range(period, len(gains)):
        avg_gain = (avg_gain * (period - 1) + gains[i]) / period
        avg_loss = (avg_loss * (period - 1) + losses[i]) / period
    if avg_loss == 0:
        return 100.0
    rs = avg_gain / avg_loss
    return round(100 - (100 / (1 + rs)), 2)


def calc_avg_volume(volumes, period=20):
    recent = [v for v in volumes[-period:] if v and v > 0]
    return sum(recent) / len(recent) if recent else None


def calc_52w_metrics(closes):
    if len(closes) < 2:
        return None, None
    high_52w = max(closes)
    current  = closes[-1]
    pct_from_high = round((high_52w - current) / high_52w * 100, 1)
    return round(high_52w, 2), pct_from_high


def extract_series(raw, ticker, field, tickers):
    series = None
    if len(tickers) > 1:
        if isinstance(raw.columns, pd.MultiIndex):
            if (field, ticker) in raw.columns:
                series = raw[(field, ticker)].dropna()
            elif (ticker, field) in raw.columns:
                series = raw[(ticker, field)].dropna()
            else:
                try:
                    series = raw.xs(ticker, axis=1, level=1)[field].dropna()
                except Exception:
                    try:
                        series = raw.xs(ticker, axis=1, level=0)[field].dropna()
                    except Exception:
                        pass
        else:
            if ticker in raw.columns:
                series = raw[ticker].dropna()
            elif field in raw.columns:
                series = raw[field].dropna()
    else:
        if field in raw.columns:
            series = raw[field].dropna()
    return series.values.tolist() if series is not None else []


# ── Load symbols from uploaded CSV bytes ───────────────────

def load_symbols_from_bytes(file_bytes):
    """Parse CSV from uploaded file bytes. Returns (symbols list, info df)."""
    df = pd.read_csv(io.BytesIO(file_bytes))
    # Accept 'Symbol' or 'symbol' column
    col_map = {c.strip().lower(): c for c in df.columns}
    if "symbol" not in col_map:
        raise ValueError(f"CSV must have a 'Symbol' column. Found: {list(df.columns)}")
    sym_col = col_map["symbol"]
    symbols = df[sym_col].dropna().astype(str).str.strip().unique().tolist()
    keep = [c for c in [sym_col, "Company Name", "Industry"] if c in df.columns]
    info = df[keep].copy().rename(columns={sym_col: "Symbol"})
    return symbols, info


# ── Load holdings from uploaded CSV bytes ──────────────────

def load_holdings_from_bytes(file_bytes):
    """
    Parse holdings CSV. Expected columns:
      Symbol, Entry Date (DD-MM-YYYY), Entry Price
    Returns list of dicts.
    """
    df = pd.read_csv(io.BytesIO(file_bytes))
    df.columns = [c.strip() for c in df.columns]
    col_map = {c.lower(): c for c in df.columns}

    if "symbol" not in col_map:
        raise ValueError(f"Holdings CSV must have a 'Symbol' column. Found: {list(df.columns)}")

    sym_col   = col_map["symbol"]
    date_col  = col_map.get("entry date", None)
    price_col = col_map.get("entry price", None)

    holdings = []
    for _, row in df.iterrows():
        sym = str(row[sym_col]).strip()
        if not sym:
            continue
        entry_date  = str(row[date_col]).strip()  if date_col  else "—"
        entry_price = row[price_col]               if price_col else None
        holdings.append({
            "symbol":      sym,
            "entry_date":  entry_date,
            "entry_price": float(entry_price) if entry_price and str(entry_price) != "nan" else None,
        })
    return holdings


# ── Exit signal check (weekly 50 EMA) ─────────────────────

def check_exit_signals(holdings, progress_callback=None):
    """
    For each holding, fetch weekly OHLC and check if current price
    is below weekly 50 EMA.
    Returns list of dicts with exit signal status for each holding.
    """
    if not holdings:
        return []

    symbols = [h["symbol"] for h in holdings]
    tickers = [s + ".NS" for s in symbols]

    if progress_callback:
        progress_callback(f"Checking exit signals for {len(symbols)} holdings (weekly data)…")

    try:
        raw = yf.download(
            tickers,
            period="2y",        # 2 years for reliable weekly 50 EMA (needs 50 weeks)
            interval="1wk",
            group_by="ticker",
            auto_adjust=True,
            progress=False,
            threads=True,
        )
    except Exception as e:
        return [{"symbol": h["symbol"], "error": str(e)} for h in holdings]

    results = []
    for holding in holdings:
        symbol = holding["symbol"]
        ticker = symbol + ".NS"
        try:
            closes = extract_series(raw, ticker, "Close", tickers)

            if not closes or len(closes) < 50:
                results.append({
                    "Symbol":       symbol,
                    "Entry Date":   holding["entry_date"],
                    "Entry Price":  holding["entry_price"],
                    "Current Price": "—",
                    "Weekly 50 EMA": "—",
                    "Signal":       "⚠ Insufficient data",
                    "P&L %":        "—",
                    "Action":       "Check manually",
                })
                continue

            current     = float(closes[-1])
            weekly_ema50 = calc_ema(closes, 50)
            exit_signal  = current < weekly_ema50

            # P&L if entry price available
            pnl_str = "—"
            if holding["entry_price"]:
                pnl = (current - holding["entry_price"]) / holding["entry_price"] * 100
                pnl_str = f"{'+' if pnl >= 0 else ''}{pnl:.1f}%"

            buffer = ((current - weekly_ema50) / weekly_ema50) * 100

            results.append({
                "Symbol":         symbol,
                "Entry Date":     holding["entry_date"],
                "Entry Price":    f"₹{holding['entry_price']:,.2f}" if holding["entry_price"] else "—",
                "Current Price":  round(current, 2),
                "Weekly 50 EMA": round(weekly_ema50, 2),
                "vs W50 EMA":    f"{'+' if buffer >= 0 else ''}{buffer:.1f}%",
                "P&L %":         pnl_str,
                "Signal":        "🚨 EXIT" if exit_signal else "✅ HOLD",
                "Action":        "Consider exiting — price below weekly 50 EMA" if exit_signal else "Hold position",
            })

        except Exception as e:
            results.append({
                "Symbol":       symbol,
                "Entry Date":   holding["entry_date"],
                "Entry Price":  holding["entry_price"],
                "Current Price": "—",
                "Weekly 50 EMA": "—",
                "Signal":       f"✗ Error: {str(e)[:50]}",
                "P&L %":        "—",
                "Action":       "Check manually",
            })

    # Sort — exits first, then holds
    results.sort(key=lambda x: 0 if "EXIT" in str(x.get("Signal","")) else 1)
    return results


# ── Download ───────────────────────────────────────────────

def download_data(symbols, progress_callback=None):
    tickers = [s + ".NS" for s in symbols]
    if progress_callback:
        progress_callback(f"Downloading data for {len(tickers)} stocks from Yahoo Finance…")
    raw = yf.download(
        tickers,
        period="1y",
        interval="1d",
        group_by="ticker",
        auto_adjust=True,
        progress=False,
        threads=True,
    )
    return raw, tickers


# ── Screener ───────────────────────────────────────────────

def run_screener(raw, symbols, tickers, progress_callback=None):
    results  = []
    rejected = []
    skipped  = []
    errors   = []

    for i, symbol in enumerate(symbols):
        ticker = symbol + ".NS"
        if progress_callback and i % 10 == 0:
            progress_callback(f"Screening {i+1}/{len(symbols)}: {symbol}")
        try:
            closes  = extract_series(raw, ticker, "Close",  tickers)
            volumes = extract_series(raw, ticker, "Volume", tickers)

            if not closes or len(closes) < 200:
                skipped.append(f"{symbol} (only {len(closes)} days of data)")
                continue

            current = float(closes[-1])
            ema50   = calc_ema(closes, 50)
            ema100  = calc_ema(closes, 100)
            ema200  = calc_ema(closes, 200)
            ema_ok  = current > ema50 and current > ema100 and current > ema200

            rsi    = calc_rsi(closes, 14)
            rsi_ok = rsi is not None and rsi > RSI_MIN

            avg_vol = calc_avg_volume(volumes, 20)
            vol_ok  = avg_vol is not None and avg_vol >= MIN_AVG_VOLUME

            high_52w, pct_from_high = calc_52w_metrics(closes)
            high_ok = pct_from_high is not None and pct_from_high <= NEAR_HIGH_PCT

            all_pass = ema_ok and rsi_ok and vol_ok and high_ok

            if all_pass:
                results.append({
                    "Symbol":          symbol,
                    "Price (₹)":       round(current, 2),
                    "50 EMA":          round(ema50,  2),
                    "100 EMA":         round(ema100, 2),
                    "200 EMA":         round(ema200, 2),
                    "vs 50 EMA":       f"+{round((current-ema50)/ema50*100,1)}%",
                    "vs 200 EMA":      f"+{round((current-ema200)/ema200*100,1)}%",
                    "RSI (14)":        rsi,
                    "Avg Vol (20d)":   f"{avg_vol/1e6:.2f}M",
                    "52W High (₹)":    high_52w,
                    "% from 52W High": f"-{pct_from_high}%",
                })
            elif ema_ok:
                rsi_str  = f"{rsi:.1f}" if rsi else "—"
                vol_str  = f"{avg_vol/1e6:.1f}M" if avg_vol else "—"
                high_str = f"-{pct_from_high}%" if pct_from_high is not None else "—"
                rejected.append({
                    "symbol": symbol,
                    "reason": f"RSI={rsi_str} Vol={vol_str} 52wH={high_str}",
                })

        except Exception as e:
            errors.append(f"{symbol}: {str(e)[:70]}")

    return results, rejected, skipped, errors


# ── Score & rank ───────────────────────────────────────────

def score_and_rank(results, universe_info):
    df = pd.DataFrame(results)
    if df.empty:
        return df

    def pct_rank(series, invert=False):
        ranked = series.rank(pct=True) * 100
        return (100 - ranked) if invert else ranked

    df["_ema200_raw"] = df["vs 200 EMA"].str.replace("+", "", regex=False).str.replace("%", "", regex=False).astype(float)
    df["_rsi_raw"]    = df["RSI (14)"].astype(float)
    df["_52w_raw"]    = df["% from 52W High"].str.replace("-", "", regex=False).str.replace("%", "", regex=False).astype(float)
    df["_vol_raw"]    = df["Avg Vol (20d)"].str.replace("M", "", regex=False).astype(float)

    df["Composite Score"] = (
        pct_rank(df["_ema200_raw"]) * W_EMA200 +
        pct_rank(df["_rsi_raw"])    * W_RSI    +
        pct_rank(df["_52w_raw"], invert=True) * W_52W +
        pct_rank(df["_vol_raw"])    * W_VOLUME
    ).round(1)

    df = df.sort_values("Composite Score", ascending=False)
    df.insert(0, "Rank", range(1, len(df) + 1))
    df = df.drop(columns=[c for c in df.columns if c.startswith("_")])

    if not universe_info.empty:
        df = df.merge(universe_info, on="Symbol", how="left")
        front = ["Rank", "Symbol"] + [c for c in ["Company Name", "Industry"] if c in df.columns]
        rest  = [c for c in df.columns if c not in front]
        df    = df[front + rest]

    return df


# ── Build Excel in memory ──────────────────────────────────

def build_excel(df, rejected, run_date, exit_signals=None):
    """Returns Excel file as bytes (in-memory, no disk write)."""
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # Sheet 1 — passed stocks
        df.to_excel(writer, index=False, sheet_name="✅ Passed")
        ws = writer.sheets["✅ Passed"]

        for i, w in enumerate([6,14,20,16,11,11,11,10,10,8,13,13,16,12,10], 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w

        hdr_fill   = PatternFill("solid", fgColor="1F4E79")
        hdr_font   = Font(color="FFFFFF", bold=True)
        alt_fill   = PatternFill("solid", fgColor="EBF5FB")
        top10_fill = PatternFill("solid", fgColor="E2EFDA")
        no_fill    = PatternFill(fill_type=None)

        headers   = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        rsi_col   = headers.index("RSI (14)")        + 1 if "RSI (14)"        in headers else None
        score_col = headers.index("Composite Score") + 1 if "Composite Score" in headers else None
        rank_col  = headers.index("Rank")            + 1 if "Rank"            in headers else None

        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx in range(2, ws.max_row + 1):
            rank_val = ws.cell(row_idx, rank_col).value if rank_col else 999
            is_top10 = isinstance(rank_val, (int, float)) and rank_val <= 10
            for cell in ws[row_idx]:
                if is_top10:
                    cell.fill = top10_fill
                elif row_idx % 2 == 0:
                    cell.fill = alt_fill
                else:
                    cell.fill = no_fill
            if rsi_col:
                rsi_cell = ws.cell(row_idx, rsi_col)
                try:
                    rval = float(rsi_cell.value)
                    rsi_cell.font = Font(color="375623" if rval >= 60 else "843C0C", bold=is_top10)
                except Exception:
                    pass
            if score_col:
                sc = ws.cell(row_idx, score_col)
                sc.font = Font(color="1F4E79", bold=True)
                sc.alignment = Alignment(horizontal="center")

        meta = ws.max_row + 2
        for label, val in [
            ("Run Date:",      run_date),
            ("Filters:",       f"Price > 50/100/200 EMA | RSI(14) > {RSI_MIN} | Avg Vol > {MIN_AVG_VOLUME:,} | Within {NEAR_HIGH_PCT}% of 52W High"),
            ("Score Weights:", f"EMA200={int(W_EMA200*100)}% RSI={int(W_RSI*100)}% 52WH={int(W_52W*100)}% Vol={int(W_VOLUME*100)}%"),
            ("Top 10:",        "Green rows = top 10 by Composite Score"),
            ("Data Source:",   "Yahoo Finance (1 year daily, auto-adjusted)"),
        ]:
            ws.cell(meta, 1, label).font = Font(bold=True)
            ws.cell(meta, 2, val)
            meta += 1

        # Sheet 2 — EMA only
        if rejected:
            rej_df = pd.DataFrame(rejected)
            rej_df.columns = ["Symbol", "Why it failed (EMA ✅, others ❌)"]
            rej_df.to_excel(writer, index=False, sheet_name="⚠ EMA Pass Only")
            ws2 = writer.sheets["⚠ EMA Pass Only"]
            ws2.column_dimensions["A"].width = 16
            ws2.column_dimensions["B"].width = 50
            for cell in ws2[1]:
                cell.fill = PatternFill("solid", fgColor="7F3F00")
                cell.font = Font(color="FFFFFF", bold=True)

        # Sheet 3 — Exit signals
        if exit_signals:
            exit_df = pd.DataFrame(exit_signals)
            exit_df.to_excel(writer, index=False, sheet_name="🚨 Exit Signals")
            ws3 = writer.sheets["🚨 Exit Signals"]

            # Column widths
            for i, w in enumerate([14, 14, 14, 14, 14, 12, 10, 14, 36], 1):
                ws3.column_dimensions[ws3.cell(1, i).column_letter].width = w

            exit_hdr_fill = PatternFill("solid", fgColor="7B0000")
            hold_fill     = PatternFill("solid", fgColor="E2EFDA")
            exit_fill     = PatternFill("solid", fgColor="FFE0E0")
            no_fill3      = PatternFill(fill_type=None)

            for cell in ws3[1]:
                cell.fill = exit_hdr_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")

            for row_idx in range(2, ws3.max_row + 1):
                signal_col = [ws3.cell(1, c).value for c in range(1, ws3.max_column + 1)]
                sig_idx    = signal_col.index("Signal") + 1 if "Signal" in signal_col else None
                sig_val    = ws3.cell(row_idx, sig_idx).value if sig_idx else ""
                is_exit    = "EXIT" in str(sig_val)

                for cell in ws3[row_idx]:
                    cell.fill = exit_fill if is_exit else (hold_fill if not is_exit else no_fill3)

                if sig_idx:
                    sig_cell = ws3.cell(row_idx, sig_idx)
                    sig_cell.font = Font(
                        color="B00000" if is_exit else "375623",
                        bold=True
                    )

    output.seek(0)
    return output.read()


# ── Main entry point called by app.py ─────────────────────

def run_full_screen(file_bytes, holdings_bytes=None, progress_callback=None):
    """
    file_bytes     : universe CSV (required) — must have 'Symbol' column
    holdings_bytes : holdings CSV (optional) — columns: Symbol, Entry Date, Entry Price
    Returns dict: { top10, all_passed, stats, exit_signals, excel_bytes, run_date }
    """
    run_date = datetime.today().strftime("%d %b %Y")

    symbols, universe_info = load_symbols_from_bytes(file_bytes)

    if progress_callback:
        progress_callback(f"Loaded {len(symbols)} symbols from CSV")

    raw, tickers = download_data(symbols, progress_callback)
    results, rejected, skipped, errors = run_screener(raw, symbols, tickers, progress_callback)

    if progress_callback:
        progress_callback("Calculating composite scores…")

    df = score_and_rank(results, universe_info)

    # ── Exit signal check ──────────────────────────────────
    exit_signals = []
    if holdings_bytes:
        try:
            holdings = load_holdings_from_bytes(holdings_bytes)
            exit_signals = check_exit_signals(holdings, progress_callback)
        except Exception as e:
            if progress_callback:
                progress_callback(f"Holdings check skipped: {str(e)}")

    excel_bytes = build_excel(df, rejected, run_date, exit_signals) if not df.empty else None

    all_passed = df.to_dict(orient="records") if not df.empty else []
    top10      = [r for r in all_passed if r.get("Rank", 99) <= 10]

    exit_count = sum(1 for e in exit_signals if "EXIT" in str(e.get("Signal", "")))

    stats = {
        "run_date":    run_date,
        "scanned":     len(symbols),
        "passed":      len(results),
        "ema_only":    len(rejected),
        "skipped":     len(skipped),
        "errors":      len(errors),
        "holdings":    len(exit_signals),
        "exit_alerts": exit_count,
    }

    return {
        "top10":        top10,
        "all_passed":   all_passed,
        "rejected":     rejected,
        "exit_signals": exit_signals,
        "stats":        stats,
        "excel_bytes":  excel_bytes,
        "run_date":     run_date,
    }
