"""
reports.py

Excel report generator for the Momentum Screener.
"""

import io

import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

import config


# ==========================================================
# Common Styles
# ==========================================================

HEADER_FILL = PatternFill(
    "solid",
    fgColor=config.HEADER_FILL,
)

HEADER_FONT = Font(
    color=config.HEADER_FONT,
    bold=True,
)

ALT_ROW_FILL = PatternFill(
    "solid",
    fgColor=config.ALT_ROW_FILL,
)

TOP10_FILL = PatternFill(
    "solid",
    fgColor=config.TOP10_FILL,
)

WARNING_FILL = PatternFill(
    "solid",
    fgColor=config.WARNING_FILL,
)

NO_FILL = PatternFill(fill_type=None)


# ==========================================================
# Helpers
# ==========================================================

def style_header(ws):

    for cell in ws[1]:

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center"
        )


def auto_width(ws):

    for column in ws.columns:

        max_length = 0

        letter = column[0].column_letter

        for cell in column:

            try:
                value = str(cell.value)

                if len(value) > max_length:
                    max_length = len(value)

            except Exception:
                pass

        ws.column_dimensions[letter].width = min(
            max_length + 3,
            40,
        )


def get_column(headers, column_name):

    if column_name not in headers:
        return None

    return headers.index(column_name) + 1


# ==========================================================
# Passed Sheet
# ==========================================================

def create_passed_sheet(
    writer,
    df,
    run_date,
):

    df.to_excel(
        writer,
        index=False,
        sheet_name=config.PASSED_SHEET,
    )

    ws = writer.sheets[
        config.PASSED_SHEET
    ]

    style_header(ws)

    auto_width(ws)

    headers = [
        ws.cell(
            1,
            c,
        ).value
        for c in range(
            1,
            ws.max_column + 1,
        )
    ]

    rank_col = get_column(
        headers,
        "Rank",
    )

    rsi_col = get_column(
        headers,
        "RSI (14)",
    )

    score_col = get_column(
        headers,
        "Composite Score",
    )

    for row in range(
        2,
        ws.max_row + 1,
    ):

        rank = (
            ws.cell(row, rank_col).value
            if rank_col
            else 999
        )

        is_top10 = (
            isinstance(rank, (int, float))
            and rank <= config.TOP_PICK_COUNT
        )

        for cell in ws[row]:

            if is_top10:

                cell.fill = TOP10_FILL

            elif row % 2 == 0:

                cell.fill = ALT_ROW_FILL

            else:

                cell.fill = NO_FILL

        if rsi_col:

            try:

                rsi = float(
                    ws.cell(
                        row,
                        rsi_col,
                    ).value
                )

                ws.cell(
                    row,
                    rsi_col,
                ).font = Font(
                    bold=is_top10,
                    color=(
                        "375623"
                        if rsi >= 60
                        else "843C0C"
                    ),
                )

            except Exception:
                pass

        if score_col:

            ws.cell(
                row,
                score_col,
            ).font = Font(
                bold=True,
                color="1F4E79",
            )

            ws.cell(
                row,
                score_col,
            ).alignment = Alignment(
                horizontal="center"
            )

    meta = ws.max_row + 2

    metadata = [

        (
            "Run Date",
            run_date,
        ),

        (
            "EMA Filter",
            f">{config.EMA_SHORT}, >{config.EMA_MEDIUM}, >{config.EMA_LONG}",
        ),

        (
            "RSI",
            f">{config.RSI_MIN}",
        ),

        (
            "Volume",
            f">{config.MIN_AVG_VOLUME:,}",
        ),

        (
            "52W High",
            f"Within {config.NEAR_HIGH_PERCENT}%",
        ),

        (
            "Score Weights",
            (
                f"EMA={int(config.WEIGHT_EMA200*100)}%  "
                f"RSI={int(config.WEIGHT_RSI*100)}%  "
                f"52W={int(config.WEIGHT_52W_HIGH*100)}%  "
                f"VOL={int(config.WEIGHT_VOLUME*100)}%"
            ),
        ),
    ]

    for label, value in metadata:

        ws.cell(
            meta,
            1,
            label,
        ).font = Font(
            bold=True
        )

        ws.cell(
            meta,
            2,
            value,
        )

        meta += 1


# ==========================================================
# EMA Only Sheet
# ==========================================================

def create_rejected_sheet(
    writer,
    rejected,
):

    if not rejected:
        return

    df = pd.DataFrame(rejected)

    df.columns = [
        "Symbol",
        "Why it failed",
    ]

    df.to_excel(
        writer,
        index=False,
        sheet_name=config.EMA_ONLY_SHEET,
    )

    ws = writer.sheets[
        config.EMA_ONLY_SHEET
    ]

    style_header(ws)

    auto_width(ws)


# ==========================================================
# Exit Sheet
# ==========================================================

def create_exit_sheet(
    writer,
    exit_signals,
):

    if not exit_signals:
        return

    df = pd.DataFrame(
        exit_signals
    )

    df.to_excel(
        writer,
        index=False,
        sheet_name="🚨 Exit Signals",
    )

    ws = writer.sheets[
        "🚨 Exit Signals"
    ]

    style_header(ws)

    auto_width(ws)

    headers = [
        ws.cell(
            1,
            c,
        ).value
        for c in range(
            1,
            ws.max_column + 1,
        )
    ]

    signal_col = get_column(
        headers,
        "Signal",
    )

    if signal_col is None:
        return

    for row in range(
        2,
        ws.max_row + 1,
    ):

        signal = str(
            ws.cell(
                row,
                signal_col,
            ).value
        )

        is_exit = "EXIT" in signal

        fill = (
            PatternFill(
                "solid",
                fgColor="FFE0E0",
            )
            if is_exit
            else TOP10_FILL
        )

        for cell in ws[row]:
            cell.fill = fill

        ws.cell(
            row,
            signal_col,
        ).font = Font(
            bold=True,
            color=(
                "B00000"
                if is_exit
                else "375623"
            ),
        )


# ==========================================================
# Main Function
# ==========================================================

def build_excel(
    passed_df,
    rejected,
    run_date,
    exit_signals=None,
):

    output = io.BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
    ) as writer:

        create_passed_sheet(
            writer,
            passed_df,
            run_date,
        )

        create_rejected_sheet(
            writer,
            rejected,
        )

        create_exit_sheet(
            writer,
            exit_signals,
        )

    output.seek(0)

    return output.read()